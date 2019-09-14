import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)

    #Acessa a aba1
    sheet = wb['Sheet1']

    #Acessa a coordenada A1 da aba 1. (ambas as funções abaixo fazem a mesma coisa)
    cell = sheet['a1']
    cell = sheet.cell(1,1)

    #Printa o valor contido na coordenada atribuída a cell
    print(cell.value)

    #Itera da segunda linha até a última linha (max_row), mas não o inclui. Então add + 1 para incluir
    for row in range(2, sheet.max_row + 1):

        #acessa todos os valores das linhas 'rows' da coluna 3
        cell = sheet.cell(row,3)

        #Multiplica o valor de todas as linhas da coluna 3 por 0.9
        corrected_price = cell.value*0.9

        #Acessa todas as linhas da coluna 4 e atribui o novo nome a essa coluna (corrected_price_cell)
        corrected_price_cell = sheet.cell(row, 4)

        #Todos os valores corrigidos serão atribuídos a todas as linhas da coluna 4
        corrected_price_cell.value = corrected_price


    #Pega as referências de qual valor quer se criar os gráficos a adiciona em values
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    #Cria as variáveis para criar gráficos
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')


    #Salva as correções feitas em um novo arquivo chamado transactions2.xlxs
    wb.save(filename)